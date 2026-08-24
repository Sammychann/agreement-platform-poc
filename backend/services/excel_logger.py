import uuid
import json
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List
import openpyxl

from config import EXCEL_LOG_PATH
from schemas.generation import AgreementFormData

HEADERS = [
    "entry_id", "timestamp", "agreement_type", "customer_name", "distributor_name",
    "location", "address", "initiator_name_and_date", "manager_name_and_date",
    "date", "equipment_json", "receiver_name", "receiver_title", "receiver_date",
    "intervet_name", "intervet_title", "intervet_date", "status",
    "customer_signature_path", "intervet_signature_path"
]

class ExcelLogger:
    def __init__(self):
        self.lock = threading.Lock()
        self.filepath = EXCEL_LOG_PATH
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        with self.lock:
            if not self.filepath.exists():
                wb = openpyxl.Workbook()
                try:
                    ws = wb.active
                    ws.title = "Log"
                    ws.append(HEADERS)
                    wb.save(self.filepath)
                finally:
                    wb.close()

    def log_entry(self, form_data: AgreementFormData, signatures_info: Dict[str, str], entry_id: Optional[str] = None) -> str:
        if not entry_id:
            entry_id = str(uuid.uuid4())
            
        timestamp = datetime.now().isoformat()
        
        # Serialize equipment list
        eq_list = []
        for eq in form_data.equipment:
            if isinstance(eq, dict):
                eq_list.append(eq)
            elif hasattr(eq, 'dict'):
                eq_list.append(eq.dict())
            else:
                eq_list.append({'equipment_name': getattr(eq, 'equipment_name', ''), 'quantity': getattr(eq, 'quantity', '')})
        equipment_json = json.dumps(eq_list)

        row = [
            entry_id,
            timestamp,
            str(form_data.agreement_type or ''),
            str(form_data.customer_name or ''),
            str(form_data.distributor_name or ''),
            str(form_data.location or ''),
            str(form_data.address or ''),
            str(form_data.initiator_name_and_date or ''),
            str(form_data.manager_name_and_date or ''),
            str(form_data.date or ''),
            equipment_json,
            str(form_data.receiver_name or ''),
            str(form_data.receiver_title or ''),
            str(form_data.receiver_date or ''),
            str(form_data.intervet_name or ''),
            str(form_data.intervet_title or ''),
            str(form_data.intervet_date or ''),
            "ACTIVE",
            str(signatures_info.get("customer_signature", "")),
            str(signatures_info.get("intervet_signature", ""))
        ]

        with self.lock:
            wb = openpyxl.load_workbook(self.filepath)
            try:
                ws = wb.active
                ws.append(row)
                wb.save(self.filepath)
            finally:
                wb.close()
            
        return entry_id

    def update_entry(self, entry_id: str, updated_fields: Dict[str, Any]) -> bool:
        with self.lock:
            if not self.filepath.exists():
                return False
                
            wb = openpyxl.load_workbook(self.filepath)
            try:
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                try:
                    id_idx = headers.index("entry_id")
                except ValueError:
                    return False
                    
                for row in ws.iter_rows(min_row=2):
                    if str(row[id_idx].value) == str(entry_id):
                        for k, v in updated_fields.items():
                            if k == 'equipment' and isinstance(v, list):
                                k = 'equipment_json'
                                v = json.dumps(v)
                            if k in headers:
                                col_idx = headers.index(k)
                                row[col_idx].value = v
                        wb.save(self.filepath)
                        return True
                return False
            finally:
                wb.close()

    def rollback_entry(self, entry_id: str) -> bool:
        with self.lock:
            if not self.filepath.exists():
                return False
                
            wb = openpyxl.load_workbook(self.filepath)
            try:
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                try:
                    id_idx = headers.index("entry_id")
                    status_idx = headers.index("status")
                except ValueError:
                    return False
                    
                for row in ws.iter_rows(min_row=2):
                    if str(row[id_idx].value) == str(entry_id):
                        row[status_idx].value = "CANCELLED"
                        wb.save(self.filepath)
                        return True
                return False
            finally:
                wb.close()

    def get_entry(self, entry_id: str) -> Optional[Dict[str, Any]]:
        with self.lock:
            if not self.filepath.exists():
                return None
                
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
            try:
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                try:
                    id_idx = headers.index("entry_id")
                except ValueError:
                    return None
                    
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[id_idx]) == str(entry_id):
                        res = dict(zip(headers, row))
                        # Parse equipment_json back to list
                        if res.get('equipment_json'):
                            try:
                                res['equipment'] = json.loads(res['equipment_json'])
                            except Exception:
                                res['equipment'] = []
                        else:
                            res['equipment'] = []
                        return res
                return None
            finally:
                wb.close()

excel_logger = ExcelLogger()
