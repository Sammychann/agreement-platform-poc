"""Agreement document validation service for batch ZIP archives."""

import zipfile
import uuid
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple
import os
import openpyxl
import pdfplumber
from docx import Document

from config import TEMP_DIR, REQUIRED_VALIDATION_FILES, FIELD_RULES
from schemas.validation import ValidationReport, CompanyValidation, FileValidation

class AgreementValidator:
    
    def validate_zip(self, zip_path: str) -> ValidationReport:
        extract_dir = TEMP_DIR / str(uuid.uuid4())
        extract_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
                
            report = self._analyze_directory(extract_dir)
            return report
            
        finally:
            self._cleanup_dir(extract_dir)
            
    def _cleanup_dir(self, directory: Path):
        if directory and directory.exists():
            shutil.rmtree(directory, ignore_errors=True)

    def _analyze_directory(self, root_dir: Path) -> ValidationReport:
        companies_data = []
        
        # Walk to find lowest level directories containing files
        for root, dirs, files in os.walk(root_dir):
            if files and not dirs:
                company_name = Path(root).name
                month = None
                
                # Check if parent is month/company structure
                parent_name = Path(root).parent.name
                if parent_name != root_dir.name and parent_name not in ['2023', '2024', '2025', '2026']:
                    company_name = parent_name
                    month = Path(root).name

                comp_val = self._validate_company_files(company_name, month, root, files)
                companies_data.append(comp_val)

        # Detect cross-company duplicates
        all_companies = [c.company_name for c in companies_data]
        dup_companies = set([x for x in all_companies if all_companies.count(x) > 1])
        
        for c in companies_data:
            if c.company_name in dup_companies:
                if c.company_name not in c.duplicates:
                    c.duplicates.append("Company appears multiple times")
        
        passed = sum(1 for c in companies_data if all(f.found and not f.missing_fields for f in c.files))
        failed = len(companies_data) - passed
        warnings = sum(len(c.duplicates) for c in companies_data)

        return ValidationReport(
            report_id=str(uuid.uuid4()),
            zip_type="auto-detected",
            total_companies=len(companies_data),
            passed=passed,
            failed=failed,
            warnings=warnings,
            companies=companies_data,
            generated_at=datetime.now()
        )
        
    def _validate_company_files(self, company_name: str, month: str, folder_path: str, files: List[str]) -> CompanyValidation:
        file_validations = []
        
        for req_file in REQUIRED_VALIDATION_FILES:
            matched_file = next((f for f in files if req_file.lower() in f.lower()), None)
            
            if not matched_file:
                file_validations.append(FileValidation(
                    filename=req_file,
                    found=False,
                    missing_fields=FIELD_RULES.get(req_file, []),
                    extra_info={}
                ))
            else:
                full_path = Path(folder_path) / matched_file
                missing_fields, extra_info = self._check_file_contents(full_path, req_file)
                
                file_validations.append(FileValidation(
                    filename=matched_file,
                    found=True,
                    missing_fields=missing_fields,
                    extra_info=extra_info
                ))
                
        return CompanyValidation(
            company_name=company_name,
            month=month,
            files=file_validations,
            duplicates=[]
        )

    def _check_file_contents(self, file_path: Path, file_type: str) -> Tuple[List[str], Dict[str, Any]]:
        required_fields = FIELD_RULES.get(file_type, [])
        ext = file_path.suffix.lower()
        content_text = ""
        missing = list(required_fields)
        extra = {}
        
        try:
            if ext == '.pdf':
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            content_text += text.lower() + " "
            elif ext == '.docx':
                doc = Document(file_path)
                for para in doc.paragraphs:
                    content_text += para.text.lower() + " "
            elif ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, data_only=True)
                try:
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            for cell in row:
                                if cell:
                                    content_text += str(cell).lower() + " "
                finally:
                    wb.close()
            
            still_missing = []
            for field in missing:
                search_term = field.replace('_', ' ').lower()
                if search_term not in content_text and field not in content_text:
                    still_missing.append(field)
            
            return still_missing, extra
            
        except Exception as e:
            return required_fields, {"error": str(e)}

    def generate_excel_report(self, report: ValidationReport) -> str:
        report_path = TEMP_DIR / f"report_{report.report_id}.xlsx"
        wb = openpyxl.Workbook()
        try:
            ws = wb.active
            ws.title = "Validation Report"
            
            headers = ["Company Name", "Month", "Status", "Missing Files", "Missing Fields", "Duplicates/Warnings"]
            ws.append(headers)
            
            for comp in report.companies:
                status = "Pass"
                missing_files = []
                missing_fields = []
                
                for f in comp.files:
                    if not f.found:
                        missing_files.append(f.filename)
                    elif f.missing_fields:
                        missing_fields.append(f"{f.filename}: {', '.join(f.missing_fields)}")
                        
                if missing_files or missing_fields:
                    status = "Fail"
                    
                ws.append([
                    comp.company_name,
                    comp.month or "",
                    status,
                    "; ".join(missing_files),
                    "; ".join(missing_fields),
                    "; ".join(comp.duplicates)
                ])
                
            wb.save(report_path)
        finally:
            wb.close()
            
        return str(report_path)

agreement_validator = AgreementValidator()
